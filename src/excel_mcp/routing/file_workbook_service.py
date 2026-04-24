"""File-backed workbook façade implementing ``RoutedWorkbookOperations`` (Epic 3).

Delegates to the same ``excel_mcp.*`` entry points as ``excel_mcp.server`` handlers
so routed calls preserve return shapes and error handling.

Debt (workbook lifecycle outside this façade): ``data.py``, ``sheet.py``,
``formatting.py``, ``workbook.py``, ``validation.py``, ``chart.py``, ``pivot.py``,
``tables.py``, and related helpers still open workbooks internally; closing them
there is deferred and out of façade scope.
"""

from __future__ import annotations

import json
import logging
from typing import Any, Dict, List, Mapping, Optional

from openpyxl import load_workbook

from excel_mcp.calculations import apply_formula as apply_formula_impl
from excel_mcp.cell_validation import get_all_validation_ranges
from excel_mcp.chart import create_chart_in_sheet as create_chart_impl
from excel_mcp.data import read_excel_range_with_metadata, write_data
from excel_mcp.exceptions import (
    CalculationError,
    ChartError,
    DataError,
    FormattingError,
    PivotError,
    SheetError,
    ValidationError,
    WorkbookError,
)
from excel_mcp.formatting import format_range as format_range_func
from excel_mcp.pivot import create_pivot_table as create_pivot_table_impl
from excel_mcp.sheet import (
    copy_range_operation,
    copy_sheet,
    delete_cols,
    delete_range_operation,
    delete_rows,
    delete_sheet,
    get_merged_ranges,
    insert_cols,
    insert_row,
    merge_range,
    rename_sheet,
    unmerge_range,
)
from excel_mcp.tables import create_excel_table as create_table_impl
from excel_mcp.validation import (
    validate_formula_in_cell_operation,
    validate_range_in_sheet_operation,
)
from excel_mcp.workbook import create_sheet as wb_create_sheet
from excel_mcp.workbook import create_workbook as wb_create_workbook
from excel_mcp.workbook import get_workbook_info

logger = logging.getLogger(__name__)


class FileWorkbookService:
    """Thin file-backed implementation of ``RoutedWorkbookOperations``."""

    def read_range_with_metadata(
        self,
        filepath: str,
        sheet_name: str,
        start_cell: str = "A1",
        end_cell: Optional[str] = None,
        preview_only: bool = False,
        *,
        operation_metadata: Optional[Mapping[str, Any]] = None,
    ) -> str:
        del preview_only, operation_metadata
        try:
            result = read_excel_range_with_metadata(
                filepath,
                sheet_name,
                start_cell,
                end_cell,
            )
            if not result or not result.get("cells"):
                return "No data found in specified range"
            return json.dumps(result, indent=2, default=str)
        except Exception as e:
            logger.error(f"Error reading data: {e}")
            raise

    def workbook_metadata(
        self,
        filepath: str,
        include_ranges: bool = False,
        *,
        operation_metadata: Optional[Mapping[str, Any]] = None,
    ) -> str:
        del operation_metadata
        try:
            result = get_workbook_info(filepath, include_ranges=include_ranges)
            return str(result)
        except WorkbookError as e:
            return f"Error: {str(e)}"
        except Exception as e:
            logger.error(f"Error getting workbook metadata: {e}")
            raise

    def read_merged_cell_ranges(
        self,
        filepath: str,
        sheet_name: str,
        *,
        operation_metadata: Optional[Mapping[str, Any]] = None,
    ) -> str:
        del operation_metadata
        try:
            return str(get_merged_ranges(filepath, sheet_name))
        except (ValidationError, SheetError) as e:
            return f"Error: {str(e)}"
        except Exception as e:
            logger.error(f"Error getting merged cells: {e}")
            raise

    def read_worksheet_data_validation(
        self,
        filepath: str,
        sheet_name: str,
        *,
        operation_metadata: Optional[Mapping[str, Any]] = None,
    ) -> str:
        del operation_metadata
        wb = load_workbook(filepath, read_only=False)
        try:
            if sheet_name not in wb.sheetnames:
                return f"Error: Sheet '{sheet_name}' not found"

            ws = wb[sheet_name]
            validations = get_all_validation_ranges(ws)

            if not validations:
                return "No data validation rules found in this worksheet"

            return json.dumps(
                {
                    "sheet_name": sheet_name,
                    "validation_rules": validations,
                },
                indent=2,
                default=str,
            )
        except Exception as e:
            logger.error(f"Error getting validation info: {e}")
            raise
        finally:
            wb.close()

    def validate_sheet_range(
        self,
        filepath: str,
        sheet_name: str,
        start_cell: str,
        end_cell: Optional[str] = None,
        *,
        operation_metadata: Optional[Mapping[str, Any]] = None,
    ) -> str:
        del operation_metadata
        try:
            range_str = start_cell if not end_cell else f"{start_cell}:{end_cell}"
            result = validate_range_in_sheet_operation(
                filepath, sheet_name, range_str
            )
            return result["message"]
        except ValidationError as e:
            return f"Error: {str(e)}"
        except Exception as e:
            logger.error(f"Error validating range: {e}")
            raise

    def validate_formula_syntax(
        self,
        filepath: str,
        sheet_name: str,
        cell: str,
        formula: str,
        *,
        operation_metadata: Optional[Mapping[str, Any]] = None,
    ) -> str:
        del operation_metadata
        try:
            result = validate_formula_in_cell_operation(
                filepath, sheet_name, cell, formula
            )
            return result["message"]
        except (ValidationError, CalculationError) as e:
            return f"Error: {str(e)}"
        except Exception as e:
            logger.error(f"Error validating formula: {e}")
            raise

    def apply_formula(
        self,
        filepath: str,
        sheet_name: str,
        cell: str,
        formula: str,
        *,
        operation_metadata: Optional[Mapping[str, Any]] = None,
    ) -> str:
        del operation_metadata
        try:
            validation = validate_formula_in_cell_operation(
                filepath, sheet_name, cell, formula
            )
            if isinstance(validation, dict) and "error" in validation:
                return f"Error: {validation['error']}"

            result = apply_formula_impl(filepath, sheet_name, cell, formula)
            return result["message"]
        except (ValidationError, CalculationError) as e:
            return f"Error: {str(e)}"
        except Exception as e:
            logger.error(f"Error applying formula: {e}")
            raise

    def format_range(
        self,
        filepath: str,
        sheet_name: str,
        start_cell: str,
        end_cell: Optional[str] = None,
        bold: bool = False,
        italic: bool = False,
        underline: bool = False,
        font_size: Optional[int] = None,
        font_color: Optional[str] = None,
        bg_color: Optional[str] = None,
        border_style: Optional[str] = None,
        border_color: Optional[str] = None,
        number_format: Optional[str] = None,
        alignment: Optional[str] = None,
        wrap_text: bool = False,
        merge_cells: bool = False,
        protection: Optional[Dict[str, Any]] = None,
        conditional_format: Optional[Dict[str, Any]] = None,
        *,
        operation_metadata: Optional[Mapping[str, Any]] = None,
    ) -> str:
        del operation_metadata
        try:
            format_range_func(
                filepath=filepath,
                sheet_name=sheet_name,
                start_cell=start_cell,
                end_cell=end_cell,
                bold=bold,
                italic=italic,
                underline=underline,
                font_size=font_size,
                font_color=font_color,
                bg_color=bg_color,
                border_style=border_style,
                border_color=border_color,
                number_format=number_format,
                alignment=alignment,
                wrap_text=wrap_text,
                merge_cells=merge_cells,
                protection=protection,
                conditional_format=conditional_format,
            )
            return "Range formatted successfully"
        except (ValidationError, FormattingError) as e:
            return f"Error: {str(e)}"
        except Exception as e:
            logger.error(f"Error formatting range: {e}")
            raise

    def write_cell_grid(
        self,
        filepath: str,
        sheet_name: str,
        data: List[List[Any]],
        start_cell: str = "A1",
        *,
        operation_metadata: Optional[Mapping[str, Any]] = None,
    ) -> str:
        del operation_metadata
        try:
            result = write_data(filepath, sheet_name, data, start_cell)
            return result["message"]
        except (ValidationError, DataError) as e:
            return f"Error: {str(e)}"
        except Exception as e:
            logger.error(f"Error writing data: {e}")
            raise

    def create_workbook(
        self,
        filepath: str,
        *,
        operation_metadata: Optional[Mapping[str, Any]] = None,
    ) -> str:
        del operation_metadata
        try:
            wb_create_workbook(filepath)
            return f"Created workbook at {filepath}"
        except WorkbookError as e:
            return f"Error: {str(e)}"
        except Exception as e:
            logger.error(f"Error creating workbook: {e}")
            raise

    def create_worksheet(
        self,
        filepath: str,
        sheet_name: str,
        *,
        operation_metadata: Optional[Mapping[str, Any]] = None,
    ) -> str:
        del operation_metadata
        try:
            result = wb_create_sheet(filepath, sheet_name)
            return result["message"]
        except (ValidationError, WorkbookError) as e:
            return f"Error: {str(e)}"
        except Exception as e:
            logger.error(f"Error creating worksheet: {e}")
            raise

    def create_chart_in_sheet(
        self,
        filepath: str,
        sheet_name: str,
        data_range: str,
        chart_type: str,
        target_cell: str,
        title: str = "",
        x_axis: str = "",
        y_axis: str = "",
        *,
        operation_metadata: Optional[Mapping[str, Any]] = None,
    ) -> str:
        del operation_metadata
        try:
            result = create_chart_impl(
                filepath=filepath,
                sheet_name=sheet_name,
                data_range=data_range,
                chart_type=chart_type,
                target_cell=target_cell,
                title=title,
                x_axis=x_axis,
                y_axis=y_axis,
            )
            return result["message"]
        except (ValidationError, ChartError) as e:
            return f"Error: {str(e)}"
        except Exception as e:
            logger.error(f"Error creating chart: {e}")
            raise

    def create_pivot_table_in_sheet(
        self,
        filepath: str,
        sheet_name: str,
        data_range: str,
        rows: List[str],
        values: List[str],
        columns: Optional[List[str]] = None,
        agg_func: str = "mean",
        *,
        operation_metadata: Optional[Mapping[str, Any]] = None,
    ) -> str:
        del operation_metadata
        try:
            result = create_pivot_table_impl(
                filepath=filepath,
                sheet_name=sheet_name,
                data_range=data_range,
                rows=rows,
                values=values,
                columns=columns or [],
                agg_func=agg_func,
            )
            return result["message"]
        except (ValidationError, PivotError) as e:
            return f"Error: {str(e)}"
        except Exception as e:
            logger.error(f"Error creating pivot table: {e}")
            raise

    def create_excel_table(
        self,
        filepath: str,
        sheet_name: str,
        data_range: str,
        table_name: Optional[str] = None,
        table_style: str = "TableStyleMedium9",
        *,
        operation_metadata: Optional[Mapping[str, Any]] = None,
    ) -> str:
        del operation_metadata
        try:
            result = create_table_impl(
                filepath=filepath,
                sheet_name=sheet_name,
                data_range=data_range,
                table_name=table_name,
                table_style=table_style,
            )
            return result["message"]
        except DataError as e:
            return f"Error: {str(e)}"
        except Exception as e:
            logger.error(f"Error creating table: {e}")
            raise

    def copy_worksheet(
        self,
        filepath: str,
        source_sheet: str,
        target_sheet: str,
        *,
        operation_metadata: Optional[Mapping[str, Any]] = None,
    ) -> str:
        del operation_metadata
        try:
            result = copy_sheet(filepath, source_sheet, target_sheet)
            return result["message"]
        except (ValidationError, SheetError) as e:
            return f"Error: {str(e)}"
        except Exception as e:
            logger.error(f"Error copying worksheet: {e}")
            raise

    def delete_worksheet(
        self,
        filepath: str,
        sheet_name: str,
        *,
        operation_metadata: Optional[Mapping[str, Any]] = None,
    ) -> str:
        del operation_metadata
        try:
            result = delete_sheet(filepath, sheet_name)
            return result["message"]
        except (ValidationError, SheetError) as e:
            return f"Error: {str(e)}"
        except Exception as e:
            logger.error(f"Error deleting worksheet: {e}")
            raise

    def rename_worksheet(
        self,
        filepath: str,
        old_name: str,
        new_name: str,
        *,
        operation_metadata: Optional[Mapping[str, Any]] = None,
    ) -> str:
        del operation_metadata
        try:
            result = rename_sheet(filepath, old_name, new_name)
            return result["message"]
        except (ValidationError, SheetError) as e:
            return f"Error: {str(e)}"
        except Exception as e:
            logger.error(f"Error renaming worksheet: {e}")
            raise

    def merge_cells(
        self,
        filepath: str,
        sheet_name: str,
        start_cell: str,
        end_cell: str,
        *,
        operation_metadata: Optional[Mapping[str, Any]] = None,
    ) -> str:
        del operation_metadata
        try:
            result = merge_range(filepath, sheet_name, start_cell, end_cell)
            return result["message"]
        except (ValidationError, SheetError) as e:
            return f"Error: {str(e)}"
        except Exception as e:
            logger.error(f"Error merging cells: {e}")
            raise

    def unmerge_cells(
        self,
        filepath: str,
        sheet_name: str,
        start_cell: str,
        end_cell: str,
        *,
        operation_metadata: Optional[Mapping[str, Any]] = None,
    ) -> str:
        del operation_metadata
        try:
            result = unmerge_range(filepath, sheet_name, start_cell, end_cell)
            return result["message"]
        except (ValidationError, SheetError) as e:
            return f"Error: {str(e)}"
        except Exception as e:
            logger.error(f"Error unmerging cells: {e}")
            raise

    def copy_cell_range(
        self,
        filepath: str,
        sheet_name: str,
        source_start: str,
        source_end: str,
        target_start: str,
        target_sheet: Optional[str] = None,
        *,
        operation_metadata: Optional[Mapping[str, Any]] = None,
    ) -> str:
        del operation_metadata
        try:
            result = copy_range_operation(
                filepath,
                sheet_name,
                source_start,
                source_end,
                target_start,
                target_sheet or sheet_name,
            )
            return result["message"]
        except (ValidationError, SheetError) as e:
            return f"Error: {str(e)}"
        except Exception as e:
            logger.error(f"Error copying range: {e}")
            raise

    def delete_cell_range(
        self,
        filepath: str,
        sheet_name: str,
        start_cell: str,
        end_cell: str,
        shift_direction: str = "up",
        *,
        operation_metadata: Optional[Mapping[str, Any]] = None,
    ) -> str:
        del operation_metadata
        try:
            result = delete_range_operation(
                filepath,
                sheet_name,
                start_cell,
                end_cell,
                shift_direction,
            )
            return result["message"]
        except (ValidationError, SheetError) as e:
            return f"Error: {str(e)}"
        except Exception as e:
            logger.error(f"Error deleting range: {e}")
            raise

    def insert_rows(
        self,
        filepath: str,
        sheet_name: str,
        start_row: int,
        count: int = 1,
        *,
        operation_metadata: Optional[Mapping[str, Any]] = None,
    ) -> str:
        del operation_metadata
        try:
            result = insert_row(filepath, sheet_name, start_row, count)
            return result["message"]
        except (ValidationError, SheetError) as e:
            return f"Error: {str(e)}"
        except Exception as e:
            logger.error(f"Error inserting rows: {e}")
            raise

    def insert_columns(
        self,
        filepath: str,
        sheet_name: str,
        start_col: int,
        count: int = 1,
        *,
        operation_metadata: Optional[Mapping[str, Any]] = None,
    ) -> str:
        del operation_metadata
        try:
            result = insert_cols(filepath, sheet_name, start_col, count)
            return result["message"]
        except (ValidationError, SheetError) as e:
            return f"Error: {str(e)}"
        except Exception as e:
            logger.error(f"Error inserting columns: {e}")
            raise

    def delete_sheet_rows(
        self,
        filepath: str,
        sheet_name: str,
        start_row: int,
        count: int = 1,
        *,
        operation_metadata: Optional[Mapping[str, Any]] = None,
    ) -> str:
        del operation_metadata
        try:
            result = delete_rows(filepath, sheet_name, start_row, count)
            return result["message"]
        except (ValidationError, SheetError) as e:
            return f"Error: {str(e)}"
        except Exception as e:
            logger.error(f"Error deleting rows: {e}")
            raise

    def delete_sheet_columns(
        self,
        filepath: str,
        sheet_name: str,
        start_col: int,
        count: int = 1,
        *,
        operation_metadata: Optional[Mapping[str, Any]] = None,
    ) -> str:
        del operation_metadata
        try:
            result = delete_cols(filepath, sheet_name, start_col, count)
            return result["message"]
        except (ValidationError, SheetError) as e:
            return f"Error: {str(e)}"
        except Exception as e:
            logger.error(f"Error deleting columns: {e}")
            raise

    def save_workbook(
        self,
        filepath: str,
        *,
        operation_metadata: Optional[Mapping[str, Any]] = None,
    ) -> str:
        del operation_metadata
        wb = load_workbook(filepath)
        try:
            wb.save(filepath)
        finally:
            wb.close()
        return f"Workbook saved: {filepath}"


__all__ = ["FileWorkbookService"]
