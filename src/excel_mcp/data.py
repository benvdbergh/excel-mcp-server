from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Dict, List, Optional
import logging
import re

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter

from excel_mcp.routing.routed_dispatch import file_backend_formula_not_evaluated_warning
from excel_mcp.routing.read_value_mode import validate_metadata_mode

from .exceptions import DataError
from .cell_utils import parse_cell_range
from .cell_validation import get_data_validation_for_cell

logger = logging.getLogger(__name__)

DEFAULT_EXPORT_MAX_ROWS = 10000


def is_xlsm_workbook(filepath: Path | str) -> bool:
    """True when ``filepath`` uses the macro-enabled ``.xlsm`` extension."""
    return str(filepath).lower().endswith(".xlsm")


def worksheet_range_has_formulas(
    ws: Worksheet,
    start_row: int,
    start_col: int,
    end_row: int,
    end_col: int,
) -> bool:
    """Return whether any cell in the rectangular range stores a formula."""
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            if cell.data_type == "f":
                return True
            val = cell.value
            if isinstance(val, str) and val.startswith("="):
                return True
    return False


def _maybe_append_file_backend_formula_warning(
    filepath: Path | str,
    ws: Worksheet,
    start_row: int,
    start_col: int,
    end_row: int,
    end_col: int,
    file_backend_warnings: List[Dict[str, str]] | None,
) -> None:
    if file_backend_warnings is None or not is_xlsm_workbook(filepath):
        return
    if worksheet_range_has_formulas(ws, start_row, start_col, end_row, end_col):
        file_backend_warnings.append(file_backend_formula_not_evaluated_warning())


def _normalize_export_max_rows(max_rows: int | None) -> int:
    if max_rows is None:
        return DEFAULT_EXPORT_MAX_ROWS
    if max_rows < 1:
        raise DataError("max_rows must be a positive integer")
    return max_rows


def _export_read_end_row(start_row: int, end_row: int, cap: int) -> int:
    """Last row to read when exporting: header row plus up to ``cap`` data rows."""
    if end_row < start_row:
        return start_row
    total_data_rows = end_row - start_row
    return start_row + min(total_data_rows, cap)


def build_worksheet_table_payload(
    sheet_name: str,
    range_str: str,
    matrix: List[List[Any]],
    *,
    max_rows: int = DEFAULT_EXPORT_MAX_ROWS,
    total_data_rows: int | None = None,
) -> Dict[str, Any]:
    """Build compact table JSON from a rectangular cell matrix (first row = headers)."""
    cap = _normalize_export_max_rows(max_rows)
    if not matrix:
        total = 0 if total_data_rows is None else max(0, total_data_rows)
        return {
            "sheet_name": sheet_name,
            "range": range_str,
            "headers": [],
            "rows": [],
            "row_count": total,
            "truncated": total > cap,
            "max_rows": cap,
        }
    headers = list(matrix[0])
    data_rows = [list(row) for row in matrix[1:]]
    total = total_data_rows if total_data_rows is not None else len(data_rows)
    truncated = total > cap
    return {
        "sheet_name": sheet_name,
        "range": range_str,
        "headers": headers,
        "rows": data_rows[:cap],
        "row_count": total,
        "truncated": truncated,
        "max_rows": cap,
    }


def _best_effort_number_format(value: int | float, fmt: str) -> str:
    """Format a numeric cell value using a subset of Excel ``number_format`` patterns."""
    fmt_part = fmt.split(";")[0]
    prefix = ""
    suffix = ""
    core = fmt_part
    for m in re.finditer(r'"([^"]*)"', fmt_part):
        literal = m.group(1)
        if m.start() < (fmt_part.find("0") if "0" in fmt_part else len(fmt_part)):
            prefix += literal
        else:
            suffix += literal
        core = core.replace(m.group(0), "")
    if "%" in core:
        pct = value * 100
        if ".00" in core or "0.00" in core:
            body = f"{pct:,.2f}" if "#,##" in core else f"{pct:.2f}"
        elif ".0" in core:
            body = f"{pct:,.1f}" if "#,##" in core else f"{pct:.1f}"
        else:
            body = f"{pct:,.0f}" if "#,##" in core else f"{pct:.0f}"
        return f"{prefix}{body}%{suffix}"
    if ".00" in core or "#,##0.00" in core:
        body = f"{value:,.2f}" if "#,##" in core else f"{value:.2f}"
    elif ".0" in core:
        body = f"{value:,.1f}" if "#,##" in core else f"{value:.1f}"
    elif "#,##" in core:
        body = f"{value:,.0f}"
    else:
        body = str(int(value)) if isinstance(value, float) and value.is_integer() else str(value)
    return f"{prefix}{body}{suffix}"


def _cell_text_value(cell: Cell) -> Any:
    """Best-effort displayed text for an openpyxl cell (weaker than Excel COM ``Range.Text``)."""
    val = cell.value
    if val is None:
        return None
    if isinstance(val, str):
        return val
    if isinstance(val, bool):
        return "TRUE" if val else "FALSE"
    if isinstance(val, (datetime, date, time)):
        return val.isoformat(sep=" ")
    fmt = cell.number_format or "General"
    if fmt == "General":
        return str(val)
    if isinstance(val, (int, float)):
        return _best_effort_number_format(val, fmt)
    return str(val)

def read_excel_range(
    filepath: Path | str,
    sheet_name: str,
    start_cell: str = "A1",
    end_cell: Optional[str] = None,
) -> List[Dict[str, Any]]:
    """Read data from Excel range."""
    try:
        wb = load_workbook(filepath, read_only=False)
        
        if sheet_name not in wb.sheetnames:
            raise DataError(f"Sheet '{sheet_name}' not found")
            
        ws = wb[sheet_name]

        # Parse start cell
        if ':' in start_cell:
            start_cell, end_cell = start_cell.split(':')
            
        # Get start coordinates
        try:
            start_coords = parse_cell_range(f"{start_cell}:{start_cell}")
            if not start_coords or not all(coord is not None for coord in start_coords[:2]):
                raise DataError(f"Invalid start cell reference: {start_cell}")
            start_row, start_col = start_coords[0], start_coords[1]
        except ValueError as e:
            raise DataError(f"Invalid start cell format: {str(e)}")

        # Determine end coordinates
        if end_cell:
            try:
                end_coords = parse_cell_range(f"{end_cell}:{end_cell}")
                if not end_coords or not all(coord is not None for coord in end_coords[:2]):
                    raise DataError(f"Invalid end cell reference: {end_cell}")
                end_row, end_col = end_coords[0], end_coords[1]
            except ValueError as e:
                raise DataError(f"Invalid end cell format: {str(e)}")
        else:
            # If no end_cell, use the full data range of the sheet
            if ws.max_row == 1 and ws.max_column == 1 and ws.cell(1, 1).value is None:
                # Handle empty sheet
                end_row, end_col = start_row, start_col
            else:
                # Use the sheet's own boundaries
                start_row, start_col = ws.min_row, ws.min_column
                end_row, end_col = ws.max_row, ws.max_column

        # Validate range bounds
        if start_row > ws.max_row or start_col > ws.max_column:
            # This case can happen if start_cell is outside the used area on a sheet with data
            # or on a completely empty sheet.
            logger.warning(
                f"Start cell {start_cell} is outside the sheet's data boundary "
                f"({get_column_letter(ws.min_column)}{ws.min_row}:{get_column_letter(ws.max_column)}{ws.max_row}). "
                f"No data will be read."
            )
            return []

        data = []
        for row in range(start_row, end_row + 1):
            row_data = []
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row, column=col)
                row_data.append(cell.value)
            if any(v is not None for v in row_data):
                data.append(row_data)

        wb.close()
        return data
    except DataError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to read Excel range: {e}")
        raise DataError(str(e))

def write_data(
    filepath: str,
    sheet_name: Optional[str],
    data: Optional[List[List]],
    start_cell: str = "A1",
) -> Dict[str, str]:
    """Write data to Excel sheet with workbook handling
    
    Headers are handled intelligently based on context.
    """
    try:
        if not data:
            raise DataError("No data provided to write")
            
        wb = load_workbook(filepath)

        # If no sheet specified, use active sheet
        if not sheet_name:
            active_sheet = wb.active
            if active_sheet is None:
                raise DataError("No active sheet found in workbook")
            sheet_name = active_sheet.title
        elif sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)

        ws = wb[sheet_name]

        # Validate start cell
        try:
            start_coords = parse_cell_range(start_cell)
            if not start_coords or not all(coord is not None for coord in start_coords[:2]):
                raise DataError(f"Invalid start cell reference: {start_cell}")
        except ValueError as e:
            raise DataError(f"Invalid start cell format: {str(e)}")

        if len(data) > 0:
            _write_data_to_worksheet(ws, data, start_cell)

        wb.save(filepath)
        wb.close()

        return {"message": f"Data written to {sheet_name}", "active_sheet": sheet_name}
    except DataError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to write data: {e}")
        raise DataError(str(e))

def _write_data_to_worksheet(
    worksheet: Worksheet, 
    data: List[List], 
    start_cell: str = "A1",
) -> None:
    """Write data to worksheet with intelligent header handling"""
    try:
        if not data:
            raise DataError("No data provided to write")

        try:
            start_coords = parse_cell_range(start_cell)
            if not start_coords or not all(x is not None for x in start_coords[:2]):
                raise DataError(f"Invalid start cell reference: {start_cell}")
            start_row, start_col = start_coords[0], start_coords[1]
        except ValueError as e:
            raise DataError(f"Invalid start cell format: {str(e)}")

        # Write data
        for i, row in enumerate(data):
            for j, val in enumerate(row):
                worksheet.cell(row=start_row + i, column=start_col + j, value=val)
    except DataError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to write worksheet data: {e}")
        raise DataError(str(e))

def read_excel_range_with_metadata(
    filepath: Path | str,
    sheet_name: str,
    start_cell: str = "A1",
    end_cell: Optional[str] = None,
    *,
    value_mode: str = "value",
    metadata_mode: str = "full",
    file_backend_warnings: List[Dict[str, str]] | None = None,
) -> Dict[str, Any]:
    """Read data from Excel range with cell metadata including validation rules.
    
    Args:
        filepath: Path to Excel file
        sheet_name: Name of worksheet
        start_cell: Starting cell address
        end_cell: Ending cell address (optional)
        value_mode: ``value`` (raw ``cell.value``) or ``text`` (best-effort display string)
        metadata_mode: ``full`` (per-cell validation metadata) or ``compact`` (omit validation)
        file_backend_warnings: Optional mutable list; when reading ``.xlsm`` with formulas,
            appends ADR 0010 ``file_backend_formula_not_evaluated`` warning entries.

    Returns:
        Dictionary containing structured cell data with metadata
    """
    metadata_mode = validate_metadata_mode(metadata_mode)
    include_validation = metadata_mode == "full"
    try:
        wb = load_workbook(filepath, read_only=False)
        
        if sheet_name not in wb.sheetnames:
            raise DataError(f"Sheet '{sheet_name}' not found")
            
        ws = wb[sheet_name]

        # Parse start cell
        if ':' in start_cell:
            start_cell, end_cell = start_cell.split(':')
            
        # Get start coordinates
        try:
            start_coords = parse_cell_range(f"{start_cell}:{start_cell}")
            if not start_coords or not all(coord is not None for coord in start_coords[:2]):
                raise DataError(f"Invalid start cell reference: {start_cell}")
            start_row, start_col = start_coords[0], start_coords[1]
        except ValueError as e:
            raise DataError(f"Invalid start cell format: {str(e)}")

        # Determine end coordinates
        if end_cell:
            try:
                end_coords = parse_cell_range(f"{end_cell}:{end_cell}")
                if not end_coords or not all(coord is not None for coord in end_coords[:2]):
                    raise DataError(f"Invalid end cell reference: {end_cell}")
                end_row, end_col = end_coords[0], end_coords[1]
            except ValueError as e:
                raise DataError(f"Invalid end cell format: {str(e)}")
        else:
            # If no end_cell, use the full data range of the sheet
            if ws.max_row == 1 and ws.max_column == 1 and ws.cell(1, 1).value is None:
                # Handle empty sheet
                end_row, end_col = start_row, start_col
            else:
                # Use the sheet's own boundaries, but respect the provided start_cell
                end_row, end_col = ws.max_row, ws.max_column
                # If start_cell is 'A1' (default), we should find the true start
                if start_cell.upper() == "A1":
                    start_row, start_col = ws.min_row, ws.min_column

        # Validate range bounds
        if start_row > ws.max_row or start_col > ws.max_column:
            # This case can happen if start_cell is outside the used area on a sheet with data
            # or on a completely empty sheet.
            logger.warning(
                f"Start cell {start_cell} is outside the sheet's data boundary "
                f"({get_column_letter(ws.min_column)}{ws.min_row}:{get_column_letter(ws.max_column)}{ws.max_row}). "
                f"No data will be read."
            )
            return {
                "range": f"{start_cell}:",
                "sheet_name": sheet_name,
                "value_mode": value_mode,
                "metadata_mode": metadata_mode,
                "cells": [],
            }

        _maybe_append_file_backend_formula_warning(
            filepath,
            ws,
            start_row,
            start_col,
            end_row,
            end_col,
            file_backend_warnings,
        )

        # Build structured cell data
        range_str = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
        range_data = {
            "range": range_str,
            "sheet_name": sheet_name,
            "value_mode": value_mode,
            "metadata_mode": metadata_mode,
            "cells": [],
        }
        
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row, column=col)
                cell_address = f"{get_column_letter(col)}{row}"
                cell_value = _cell_text_value(cell) if value_mode == "text" else cell.value
                
                cell_data = {
                    "address": cell_address,
                    "value": cell_value,
                    "row": row,
                    "column": col
                }
                
                # Add validation metadata if requested
                if include_validation:
                    validation_info = get_data_validation_for_cell(ws, cell_address)
                    if validation_info:
                        cell_data["validation"] = validation_info
                    else:
                        cell_data["validation"] = {"has_validation": False}
                
                range_data["cells"].append(cell_data)

        wb.close()
        return range_data
        
    except DataError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to read Excel range with metadata: {e}")
        raise DataError(str(e))


def export_excel_worksheet_table(
    filepath: Path | str,
    sheet_name: str,
    start_cell: str = "A1",
    end_cell: Optional[str] = None,
    max_rows: int = DEFAULT_EXPORT_MAX_ROWS,
) -> Dict[str, Any]:
    """Read worksheet range as a compact table (first row headers, rest data rows)."""
    cap = _normalize_export_max_rows(max_rows)
    try:
        wb = load_workbook(filepath, read_only=False)

        if sheet_name not in wb.sheetnames:
            raise DataError(f"Sheet '{sheet_name}' not found")

        ws = wb[sheet_name]

        raw_start = start_cell
        ec = end_cell
        if ":" in raw_start:
            parts = raw_start.split(":", 1)
            raw_start, ec = parts[0].strip(), parts[1].strip() if ec is None else ec

        try:
            start_coords = parse_cell_range(f"{raw_start}:{raw_start}")
            if not start_coords or not all(coord is not None for coord in start_coords[:2]):
                raise DataError(f"Invalid start cell reference: {raw_start}")
            start_row, start_col = start_coords[0], start_coords[1]
        except ValueError as e:
            raise DataError(f"Invalid start cell format: {str(e)}")

        if ec:
            try:
                end_coords = parse_cell_range(f"{ec}:{ec}")
                if not end_coords or not all(coord is not None for coord in end_coords[:2]):
                    raise DataError(f"Invalid end cell reference: {ec}")
                end_row, end_col = end_coords[0], end_coords[1]
            except ValueError as e:
                raise DataError(f"Invalid end cell format: {str(e)}")
        else:
            if ws.max_row == 1 and ws.max_column == 1 and ws.cell(1, 1).value is None:
                end_row, end_col = start_row, start_col
            else:
                end_row, end_col = ws.max_row, ws.max_column
                if raw_start.upper() == "A1":
                    start_row, start_col = ws.min_row, ws.min_column

        if start_row > ws.max_row or start_col > ws.max_column:
            range_str = f"{get_column_letter(start_col)}{start_row}:"
            wb.close()
            return build_worksheet_table_payload(
                sheet_name, range_str, [], max_rows=cap
            )

        range_str = (
            f"{get_column_letter(start_col)}{start_row}:"
            f"{get_column_letter(end_col)}{end_row}"
        )
        read_end_row = _export_read_end_row(start_row, end_row, cap)
        total_data_rows = max(0, end_row - start_row)
        matrix: List[List[Any]] = []
        for row in range(start_row, read_end_row + 1):
            matrix.append(
                [ws.cell(row=row, column=col).value for col in range(start_col, end_col + 1)]
            )

        wb.close()
        return build_worksheet_table_payload(
            sheet_name,
            range_str,
            matrix,
            max_rows=cap,
            total_data_rows=total_data_rows,
        )

    except DataError:
        raise
    except Exception as e:
        logger.error(f"Failed to export worksheet table: {e}")
        raise DataError(str(e))
