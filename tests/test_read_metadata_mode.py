"""BEN-136: metadata_mode full vs compact on read_range_with_metadata."""

import json
import os
import sys

import pytest
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation

_REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_SRC = os.path.join(_REPO_ROOT, "src")
if _SRC not in sys.path:
    sys.path.insert(0, _SRC)

from excel_mcp.routing.file_workbook_service import FileWorkbookService  # noqa: E402
from excel_mcp.routing.read_value_mode import (  # noqa: E402
    validate_metadata_mode,
)
from excel_mcp.data import read_excel_range_with_metadata  # noqa: E402


def _workbook_with_list_validation(tmp_path):
    p = tmp_path / "validated.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "yes"
    dv = DataValidation(type="list", formula1='"yes,no,maybe"', allow_blank=True)
    dv.add(ws["A1"])
    ws.add_data_validation(dv)
    wb.save(p)
    return str(p.resolve())


def test_validate_metadata_mode_rejects_unknown() -> None:
    with pytest.raises(ValueError, match="Invalid metadata_mode"):
        validate_metadata_mode("minimal")


def test_read_excel_range_with_metadata_full_includes_validation(tmp_path) -> None:
    path = _workbook_with_list_validation(tmp_path)
    data = read_excel_range_with_metadata(path, "Sheet1", "A1", "A1", metadata_mode="full")
    assert data["metadata_mode"] == "full"
    cell = data["cells"][0]
    assert "validation" in cell
    assert cell["validation"]["has_validation"] is True
    assert cell["validation"]["validation_type"] == "list"


def test_read_excel_range_with_metadata_compact_omits_validation(tmp_path) -> None:
    path = _workbook_with_list_validation(tmp_path)
    data = read_excel_range_with_metadata(path, "Sheet1", "A1", "A1", metadata_mode="compact")
    assert data["metadata_mode"] == "compact"
    cell = data["cells"][0]
    assert "validation" not in cell
    assert set(cell.keys()) == {"address", "value", "row", "column"}


def test_read_excel_range_with_metadata_default_is_full(tmp_path) -> None:
    path = _workbook_with_list_validation(tmp_path)
    data = read_excel_range_with_metadata(path, "Sheet1", "A1", "A1")
    assert data["metadata_mode"] == "full"
    assert "validation" in data["cells"][0]


def test_file_workbook_service_passes_metadata_mode(tmp_path) -> None:
    path = _workbook_with_list_validation(tmp_path)
    svc = FileWorkbookService()
    compact = json.loads(
        svc.read_range_with_metadata(path, "Sheet1", "A1", "A1", metadata_mode="compact")
    )
    full = json.loads(
        svc.read_range_with_metadata(path, "Sheet1", "A1", "A1", metadata_mode="full")
    )
    assert compact["metadata_mode"] == "compact"
    assert "validation" not in compact["cells"][0]
    assert full["metadata_mode"] == "full"
    assert "validation" in full["cells"][0]


def test_file_workbook_service_invalid_metadata_mode() -> None:
    svc = FileWorkbookService()
    with pytest.raises(ValueError, match="Invalid metadata_mode"):
        svc.read_range_with_metadata("/abs/b.xlsx", "S", metadata_mode="sparse")
