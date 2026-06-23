"""ADR 0008 lifecycle tools and ADR 0009 discovery (COM-gated)."""

from __future__ import annotations

import json
import os
import sys
from pathlib import Path

import pytest
from openpyxl import Workbook

_REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_SRC = os.path.join(_REPO_ROOT, "src")
if _SRC not in sys.path:
    sys.path.insert(0, _SRC)


def test_excel_open_workbook_when_com_unavailable_returns_error(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    import excel_mcp.server as srv

    p = tmp_path / "x.xlsx"
    Workbook().save(p)
    monkeypatch.setitem(srv.__dict__, "_COM_WORKBOOK_SERVICE", None)
    out = srv.excel_open_workbook(str(p.resolve()))
    assert "not available" in out.lower()


def test_excel_list_open_workbooks_when_com_unavailable_returns_error(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    import excel_mcp.server as srv

    monkeypatch.setitem(srv.__dict__, "_COM_WORKBOOK_SERVICE", None)
    out = srv.excel_list_open_workbooks()
    assert "not available" in out.lower()


def test_excel_list_open_workbooks_passes_detail_to_service(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    import excel_mcp.server as srv

    class _StubCom:
        def list_open_workbooks(self, detail=None):
            return json.dumps({"detail": detail, "workbooks": []})

    monkeypatch.setitem(srv.__dict__, "_COM_WORKBOOK_SERVICE", _StubCom())
    out = srv.excel_list_open_workbooks(detail="active_context")
    assert json.loads(out)["detail"] == "active_context"


def test_excel_close_workbook_when_com_unavailable_returns_error(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    import excel_mcp.server as srv

    p = tmp_path / "y.xlsx"
    Workbook().save(p)
    monkeypatch.setitem(srv.__dict__, "_COM_WORKBOOK_SERVICE", None)
    out = srv.excel_close_workbook(str(p.resolve()), save=False)
    assert "not available" in out.lower()


def test_evaluate_range_when_com_unavailable_returns_error(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    import excel_mcp.server as srv

    p = tmp_path / "recalc.xlsx"
    Workbook().save(p)
    monkeypatch.setitem(srv.__dict__, "_COM_WORKBOOK_SERVICE", None)
    out = srv.evaluate_range(str(p.resolve()), "Sheet")
    assert "not available" in out.lower()


def test_evaluate_range_rejects_file_transport(tmp_path: Path) -> None:
    import excel_mcp.server as srv

    p = tmp_path / "recalc.xlsx"
    Workbook().save(p)
    out = srv.evaluate_range(
        str(p.resolve()), "Sheet", workbook_transport="file"
    )
    assert out.startswith("Error:")
    assert "workbook_transport=file" in out


def test_create_workbook_open_in_excel_notes_when_com_unavailable(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    import excel_mcp.server as srv

    p = tmp_path / "created.xlsx"
    path = str(p.resolve())
    monkeypatch.setitem(srv.__dict__, "_COM_WORKBOOK_SERVICE", None)
    out = srv.create_workbook(path, workbook_transport="file", open_in_excel=True)
    assert p.is_file()
    assert "open_in_excel" in out.lower() or "ignored" in out.lower()
