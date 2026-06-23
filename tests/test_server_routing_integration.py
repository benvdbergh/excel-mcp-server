"""Integration-style routing through server handlers (Epic 5 Story 5-3)."""

from __future__ import annotations

import json
import logging
from pathlib import Path

import pytest
from openpyxl import Workbook


def test_read_path_logs_routing_json(
    tmp_path: Path,
    caplog: pytest.LogCaptureFixture,
) -> None:
    caplog.set_level(logging.INFO, logger="excel-mcp.routing")
    p = tmp_path / "route_read.xlsx"
    Workbook().save(p)
    path = str(p.resolve())

    from excel_mcp import server as srv

    out = srv.get_workbook_metadata(path, workbook_transport="file")
    assert "Sheet" in out or "Error" not in out[:20]

    payloads = []
    for rec in caplog.records:
        if rec.name == "excel-mcp.routing" and rec.levelno == logging.INFO:
            try:
                payloads.append(json.loads(rec.getMessage()))
            except json.JSONDecodeError:
                continue
    assert payloads, "expected JSON routing log"
    last = payloads[-1]
    assert last["workbook_transport"] == "file"
    assert last["workbook_backend"] == "file"
    assert last["routing_reason"] == "forced_file"
    assert last["operation_name"] == "workbook_metadata"
    assert last["mcp_tool_name"] == "get_workbook_metadata"


def test_read_data_from_excel_invalid_value_mode(tmp_path: Path) -> None:
    p = tmp_path / "route_read.xlsx"
    Workbook().save(p)
    path = str(p.resolve())

    from excel_mcp import server as srv

    out = srv.read_data_from_excel(path, "Sheet", value_mode="raw")
    assert out.startswith("Error:")
    assert "Invalid value_mode" in out


def test_read_data_from_excel_value_mode_file_backend(tmp_path: Path) -> None:
    p = tmp_path / "route_read_fmt.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet"
    ws["A1"] = 100
    ws["A1"].number_format = "0.00"
    wb.save(p)
    path = str(p.resolve())

    from excel_mcp import server as srv

    out = srv.read_data_from_excel(
        path, "Sheet", "A1", "A1", workbook_transport="file", value_mode="text"
    )
    data = json.loads(out)
    assert data["value_mode"] == "text"
    assert data["cells"][0]["value"] == "100.00"


def test_read_data_from_excel_xlsm_formula_warning_in_envelope(tmp_path: Path) -> None:
    p = tmp_path / "formula.xlsm"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet"
    ws["A1"] = "=2+2"
    wb.save(p)
    path = str(p.resolve())

    from excel_mcp import server as srv
    from excel_mcp.routing.routed_dispatch import (
        FILE_BACKEND_FORMULA_NOT_EVALUATED_CODE,
    )

    out = srv.read_data_from_excel(
        path,
        "Sheet",
        "A1",
        "A1",
        workbook_transport="file",
        include_routing_metadata=True,
    )
    envelope = json.loads(out)
    assert envelope["_meta"]["workbook_backend"] == "file"
    assert len(envelope["warnings"]) == 1
    assert envelope["warnings"][0]["code"] == FILE_BACKEND_FORMULA_NOT_EVALUATED_CODE
    assert envelope["result"]["cells"][0]["address"] == "A1"


def test_read_data_from_excel_xlsm_formula_legacy_response_unchanged(tmp_path: Path) -> None:
    p = tmp_path / "formula.xlsm"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet"
    ws["A1"] = "=2+2"
    wb.save(p)
    path = str(p.resolve())

    from excel_mcp import server as srv

    out = srv.read_data_from_excel(
        path,
        "Sheet",
        "A1",
        "A1",
        workbook_transport="file",
        include_routing_metadata=False,
    )
    data = json.loads(out)
    assert "warnings" not in data
    assert "result" not in data
    assert data["cells"][0]["address"] == "A1"


def test_export_worksheet_table_file_routing(tmp_path: Path) -> None:
    p = tmp_path / "route_export.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet"
    ws.append(["Name", "Qty"])
    ws.append(["item", 3])
    wb.save(p)
    path = str(p.resolve())

    from excel_mcp import server as srv

    out = srv.export_worksheet_table(path, "Sheet", workbook_transport="file")
    data = json.loads(out)
    assert data["headers"] == ["Name", "Qty"]
    assert data["rows"] == [["item", 3]]
    assert data["row_count"] == 1


def test_write_path_file_backend(tmp_path: Path) -> None:
    p = tmp_path / "route_write.xlsx"
    Workbook().save(p)
    path = str(p.resolve())

    from excel_mcp import server as srv

    msg = srv.write_data_to_excel(
        path,
        "Sheet",
        [["x"]],
        start_cell="A1",
        workbook_transport="file",
    )
    assert "success" in msg.lower() or "written" in msg.lower() or "Error" not in msg
